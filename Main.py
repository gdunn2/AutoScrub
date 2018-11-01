import os
import openpyxl as opxl

from Scrub import scrub

#variables
maxPrice = 400000
minTime = 365 #days since sold
sold = "Sold:"

inheritedPath = "New/Inherited/"

wb = opxl.load_workbook( inheritedPath + 'JDunn12Jul.xlsx')
list = wb.get_active_sheet()

for p in range(1,list.max_row):

    address = list.cell(row = p, column = 2).value
    city = list.cell(row = p, column = 3).value
    zip = list.cell(row = p, column = 5).value

    print(scrub(address, zip, maxPrice, minTime, sold))
